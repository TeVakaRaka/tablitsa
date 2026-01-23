from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from ..models import Table


class XLSXExporter:
    """Export tables to Excel XLSX format."""

    def __init__(self):
        self.header_font = Font(bold=True)
        self.header_fill = PatternFill(
            start_color="E0E0E0",
            end_color="E0E0E0",
            fill_type="solid"
        )
        self.header_alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
        self.cell_alignment = Alignment(
            vertical="center",
            wrap_text=True
        )
        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

    def export(
        self,
        tables: List[Table],
        output_path: str,
        include_headers: bool = True,
        skip_empty: bool = False,
    ):
        """Export tables to XLSX file."""
        workbook = Workbook()

        # Remove default sheet
        if workbook.active:
            workbook.remove(workbook.active)

        for table in tables:
            self._add_table_sheet(
                workbook, table,
                include_headers=include_headers,
                skip_empty=skip_empty
            )

        # If no tables, create empty sheet
        if not workbook.worksheets:
            workbook.create_sheet("Empty")

        workbook.save(output_path)

    def _add_table_sheet(
        self,
        workbook: Workbook,
        table: Table,
        include_headers: bool,
        skip_empty: bool,
    ):
        """Add a table as a worksheet."""
        # Create sheet with valid name
        sheet_name = self._sanitize_sheet_name(table.name)
        sheet = workbook.create_sheet(title=sheet_name)

        # Get enabled columns
        enabled_cols = [col for col in table.columns if col.enabled]
        if not enabled_cols:
            return

        current_row = 1

        # Write headers
        if include_headers:
            for col_idx, col in enumerate(enabled_cols, 1):
                cell = sheet.cell(row=current_row, column=col_idx)
                cell.value = col.name
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.header_alignment
                cell.border = self.thin_border
            current_row += 1

        # Write data
        data_matrix = table.get_data_matrix(enabled_only=True)

        for row_data in data_matrix:
            # Skip empty rows if requested
            if skip_empty and all(not cell.strip() for cell in row_data):
                continue

            for col_idx, value in enumerate(row_data, 1):
                cell = sheet.cell(row=current_row, column=col_idx)
                cell.value = value
                cell.alignment = self.cell_alignment
                cell.border = self.thin_border

            current_row += 1

        # Auto-adjust column widths
        for col_idx, col in enumerate(enabled_cols, 1):
            column_letter = get_column_letter(col_idx)
            max_length = len(col.name)

            for row in range(2 if include_headers else 1, current_row):
                cell_value = sheet.cell(row=row, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))

            # Set width with some padding
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width

    def _sanitize_sheet_name(self, name: str) -> str:
        """Sanitize sheet name for Excel compatibility."""
        # Excel sheet names have restrictions
        invalid_chars = ["\\", "/", "*", "?", ":", "[", "]"]
        result = name

        for char in invalid_chars:
            result = result.replace(char, "_")

        # Max 31 characters
        if len(result) > 31:
            result = result[:31]

        # Cannot be empty
        if not result:
            result = "Sheet"

        return result
